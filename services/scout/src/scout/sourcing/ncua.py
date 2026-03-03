"""NCUA Credit Unions — sourcing federally insured credit unions from bulk ZIP download."""

import csv
import io
import logging
import tempfile
import zipfile
from datetime import date

import httpx
from openpyxl import load_workbook

from scout.sourcing.base import CompanyRecord, CompanySource

log = logging.getLogger("scout.sourcing.ncua")

_URL_TEMPLATE = (
    "https://www.ncua.gov/files/publications/analysis/"
    "federally-insured-credit-union-list-{month}-{year}.zip"
)

_MONTHS = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
]

_REQUIRED_COLUMNS = {"Charter number", "Credit Union name"}


class NcuaSource(CompanySource):
    """Fetch federally insured credit unions from NCUA bulk ZIP/CSV download."""

    name = "ncua"

    def fetch(self, max_records: int = 0) -> list[CompanyRecord]:
        url = self._discover_url()
        if not url:
            log.error("NCUA: could not find a valid ZIP download URL")
            return []

        log.info("NCUA: downloading from %s", url)
        csv_text = self._download_and_extract(url)
        if not csv_text:
            return []

        records = self._parse_csv(csv_text, max_records)
        log.info("NCUA: fetched %d credit unions", len(records))
        return records

    def _discover_url(self) -> str | None:
        """Try current month backwards (up to 12 months) with HEAD requests."""
        today = date.today()
        year = today.year
        month_idx = today.month  # 1-based

        for _ in range(12):
            month_name = _MONTHS[month_idx - 1]
            url = _URL_TEMPLATE.format(month=month_name, year=year)

            try:
                resp = httpx.head(url, timeout=10.0, follow_redirects=True)
                if resp.status_code == 200:
                    log.info("NCUA: found ZIP at %s", url)
                    return url
            except httpx.HTTPError:
                pass

            month_idx -= 1
            if month_idx == 0:
                month_idx = 12
                year -= 1

        return None

    def _download_and_extract(self, url: str) -> str | None:
        """Download ZIP, extract CSV or XLSX, return contents as CSV string."""
        try:
            resp = httpx.get(url, timeout=120.0, follow_redirects=True)
            resp.raise_for_status()
        except httpx.HTTPError as exc:
            log.error("NCUA: download failed: %s", exc)
            return None

        try:
            with tempfile.SpooledTemporaryFile(max_size=50 * 1024 * 1024) as tmp:
                tmp.write(resp.content)
                tmp.seek(0)

                with zipfile.ZipFile(tmp) as zf:
                    names = zf.namelist()
                    csv_names = [n for n in names if n.lower().endswith(".csv")]
                    xlsx_names = [n for n in names if n.lower().endswith(".xlsx")]

                    if csv_names:
                        target = csv_names[0]
                        log.info("NCUA: extracting CSV %s from ZIP", target)
                        raw = zf.read(target)
                        return raw.decode("utf-8", errors="replace")

                    if xlsx_names:
                        target = xlsx_names[0]
                        log.info("NCUA: extracting XLSX %s from ZIP", target)
                        return self._xlsx_to_csv(zf.read(target))

                    log.error("NCUA: no CSV or XLSX file in ZIP (contents: %s)", names)
                    return None

        except (zipfile.BadZipFile, OSError) as exc:
            log.error("NCUA: failed to process ZIP: %s", exc)
            return None

    @staticmethod
    def _xlsx_to_csv(data: bytes) -> str | None:
        """Convert XLSX bytes to CSV string, normalising header names."""
        try:
            wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return None

            buf = io.StringIO()
            writer = csv.writer(buf)
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    # Normalise headers: collapse newlines/whitespace
                    row = tuple(
                        " ".join(str(c).split()).strip() if c is not None else ""
                        for c in row
                    )
                    first = False
                writer.writerow(row)
            wb.close()
            return buf.getvalue()
        except Exception as exc:
            log.error("NCUA: failed to convert XLSX to CSV: %s", exc)
            return None

    def _parse_csv(self, csv_text: str, max_records: int) -> list[CompanyRecord]:
        """Parse CSV text into CompanyRecord list, dedup by CU_NUMBER."""
        reader = csv.DictReader(io.StringIO(csv_text))

        if reader.fieldnames is None:
            log.error("NCUA: CSV has no header row")
            return []

        available = set(reader.fieldnames)
        missing = _REQUIRED_COLUMNS - available
        if missing:
            log.error("NCUA: CSV missing required columns %s (available: %s)", missing, sorted(available))
            return []

        records: list[CompanyRecord] = []
        seen: set[str] = set()

        for row in reader:
            if max_records > 0 and len(records) >= max_records:
                break

            cu_number = (row.get("Charter number") or "").strip()
            cu_name = (row.get("Credit Union name") or "").strip()
            if not cu_number or not cu_name or cu_number in seen:
                continue
            seen.add(cu_number)

            total_assets = None
            raw_assets = (row.get("Total assets") or "").strip()
            if raw_assets:
                try:
                    total_assets = int(float(raw_assets))
                except ValueError:
                    pass

            records.append(CompanyRecord(
                name=cu_name.title(),
                source="ncua",
                source_id=cu_number,
                city=((row.get("City (Mailing address)") or "").strip().title()) or None,
                state=((row.get("State (Mailing address)") or "").strip().upper()) or None,
                total_assets=total_assets,
                industry="Credit Union",
            ))

        return records
